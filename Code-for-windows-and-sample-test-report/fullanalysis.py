import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pathlib import Path
from collections import Counter
from datetime import datetime

suspicious_terms = [
    "login", "admin", "backup", "password", "private", "secret", "confidential",
    "config", "db", "database", "test", "old", "hidden", "user", "account",
    "auth", "restricted", "archive", "key", "token",
    "phone", "email", "contact", "whatsapp", "facebook", "linkedin", "@gmail.com", "@yahoo.com", "gmail",
    "+1", "+44", "+91", "telegram", "discord", "skype", "twitter.com", "t.me/", "instagram", "+64",
    "wallet address", "metamask", "usdt", "keylogger", "token grabber", "exe", "malware", "stealer",
    "ransomware", "shell", "backdoor", "payload", "exploit", "beef", "hook", "c2", "dll", "kernel",
    "daemon", "ip", "mac", "command and control", "dropper", "autoexec", "autorun", "shellcode",
    "system()", "exec(", "cmd", "cmd.exe", "powershell", "registry key", "validate", "credentials",
    "username", "password", "redirect", "2FA bypass", "session token", "auth token", "otp",
    "admin panel", "admin.php", "cpanel", "dashboard", "ftp", "webmail", "live support", "eval(",
    "fetch(", "xhr.open(", "window.location", "innerHTML", "onload", "cloudflare", "ngrok", "firebaseio",
    "herokuapp", "freehosting", "namecheap", "freenom", "godaddy", "hostinger", "wordpress.com",
    "blogspot.com", "base64_decode", "atob(", "btoa(", "unescape(", "obfuscate", "pack", "encrypt",
    "xor", "charCodeAt", "admin login", "panel access", "manage users", "credentials.txt", "logs",
    "session.log", "access.log", "transactions", "money", "bank transfer", "ifsc", "iban",
    "routing number", "swift code", "gateway", "broker ID", "account", "deposit address",
    "wallet recovery", "mnemonic phrase", "private key", "seed phrase", "api_key", "client_id",
    "secret_key", "token=", "api.secret", "jwt", "access_token", "env", ".env", "db_password",
    "db_user", "db_host", "config.php", "wp-config.php", ".env.production", ".env.local", "ftp_password",
    "smtp_password", "email_password", "smtp", "mailgun", "sendgrid", "smtp server", "webhook",
    "webhook.site", "endpoint", "callback", "callback_url", "payment processor", "payu", "stripe",
    "razorpay", "paypal", "skrill", "perfectmoney", "payeer", "payoneer", "btc", "eth", "xrp",
    "ltc", "usdc", "usdt", "erc20", "trc20", "bep20", "withdrawal fee", "deposit fee", "fake reviews",
    "testimonials", "as seen on", "aml", "fma warning", "iosco alert", "asx warning", "fca warning",
    "sec alert", "asic warning", "baFin warning", "amf warning", "ico", "whitelist", "kyc document",
    "passport", "national id", "proof of address", "proof of funds", "bank statement", "fraudulent",
    "fraud warning", "unlicensed", "unregulated", "scam alert", "cease and desist", "legal notice",
    "arbitration", "class action", "chargeback", "dispute", "complaint", "victim", "victim support",
    "dev@", "admin@", "webmaster@", "info@", "support@", "hostmaster@", "abuse@", "sales@", "noreply@",
    "no-reply@", "team@", "protonmail", "tutanota", "outlook", "viber", "wechat", "signal", "line.me",
    "snapchat", "kik", "email: ", "phone: ", "contact:", "telegram:", "discord:", "twitter:", "github",
    "bitbucket", "gitlab", "whois", "founder", "ceo", "developer", "engineer", "project lead",
    "contact us", "staff", "support team", "company registration", "incorporation", "business number",
    "abn", "vat", "gst", "country of incorporation", "uk companies house", "finra", "sec.gov",
    "open corporates", "company address", "office address", "address:", "eval(", "crypto-miner",
    "cryptojacker", "web skimmer", "magecart", "formgrabber", "keystroke", "window.open(",
    "document.cookie", "fetch(", "axios.post(", "axios.get(", "jquery.post(", "jquery.get(",
    "new XMLHttpRequest", "navigator.sendBeacon", "postMessage(", "open in new tab", "base64,",
    "xor_encrypt", "hex_encode", "percent_encode", "shell_exec(", "phpinfo(", "include(", "require(",
    "passthru(", "popen(", "proc_open(", "curl_exec(", "fsock", "password", "username", "login",
    "user:", "pass:", "cred=", "user_id", "login_id", "passwd", "pwd", "auth", "token", "session_id",
    "session_token", "bearer", "private_key", "public_key", "secret", "key=", "api_key", "api_secret",
    "ssh", "ssh-rsa", "-----BEGIN RSA PRIVATE KEY-----", "-----BEGIN OPENSSH PRIVATE KEY-----",
    "-----BEGIN DSA PRIVATE KEY-----", "-----BEGIN EC PRIVATE KEY-----", "-----BEGIN CERTIFICATE-----",
    "ssn", "dob", "date of birth", "credit card", "ccv", "cvv", "expiry", "pan", "cardholder",
    "billing address", "sort code", "routing number", "iban", "bic", "swift"
]

def ensure_analysis_folder(base_folder):
    analysis_dir = Path(base_folder) / "Analysis"
    analysis_dir.mkdir(exist_ok=True)
    return analysis_dir

def update_file_type_stats(root_folder, analysis_dir):
    # Count and total size by file type
    ext_counter = Counter()
    size_by_ext = Counter()
    for dirpath, _, filenames in os.walk(root_folder):
        if "Analysis" in dirpath:
            continue  # Skip Analysis folder
        for f in filenames:
            ext = Path(f).suffix.lower()
            ext_key = ext.lstrip('.').upper() if ext else 'NO EXT'
            full_path = os.path.join(dirpath, f)
            try:
                size = os.path.getsize(full_path)
            except OSError:
                size = 0
            ext_counter[ext_key] += 1
            size_by_ext[ext_key] += size

    today = datetime.now().strftime("%Y-%m-%d")
    types_sorted = sorted(ext_counter.keys())
    df_new = pd.DataFrame({
        "File Type": types_sorted,
        f"{today} (Count)": [ext_counter[t] for t in types_sorted],
        f"{today} (Size MB)": [f"{size_by_ext[t]/1024/1024:.2f}" for t in types_sorted]
    })

    outpath = analysis_dir / "file_type_stats.csv"
    if outpath.exists():
        df_old = pd.read_csv(outpath)
        df_all = pd.merge(df_old, df_new, on="File Type", how="outer").sort_values("File Type").reset_index(drop=True)
        df_all.fillna("", inplace=True)
    else:
        df_all = df_new
    df_all.to_csv(outpath, index=False)
    print(f"File type stats updated: {outpath}")

def update_top10_largest_files(root_folder, analysis_dir):
    files_info = []
    for dirpath, _, filenames in os.walk(root_folder):
        if "Analysis" in dirpath:
            continue  # Skip Analysis folder itself
        for f in filenames:
            full_path = os.path.join(dirpath, f)
            rel_path = os.path.relpath(full_path, root_folder)
            ext = Path(f).suffix.lower().lstrip('.').upper() if Path(f).suffix else 'NO EXT'
            try:
                size = os.path.getsize(full_path)
            except OSError:
                size = 0
            files_info.append({
                "File Name": f,
                "Relative Path": rel_path,
                "File Type": ext,
                "Size (MB)": size / 1024 / 1024
            })
    files_info = sorted(files_info, key=lambda x: x["Size (MB)"], reverse=True)[:10]

    today = datetime.now().strftime("%Y-%m-%d")
    df_new = pd.DataFrame(files_info)
    df_new.rename(columns={
        "Size (MB)": f"{today} Size (MB)"
    }, inplace=True)

    outpath = analysis_dir / "top10_largest_files.csv"
    if outpath.exists():
        df_old = pd.read_csv(outpath)
        merge_cols = ["File Name", "Relative Path", "File Type"]
        df_all = pd.merge(df_old, df_new, on=merge_cols, how="outer")
        df_all = df_all.sort_values(by=merge_cols).reset_index(drop=True)
        df_all.fillna("", inplace=True)
    else:
        df_all = df_new
    df_all.to_csv(outpath, index=False)
    print(f"Top 10 largest files updated: {outpath}")
def update_suspicious_files_folders(root_folder, analysis_dir):
    matches = {term: [] for term in suspicious_terms}
    for dirpath, dirnames, filenames in os.walk(root_folder):
        if "Analysis" in dirpath:
            continue
        # Check folders
        for d in dirnames:
            for term in suspicious_terms:
                if term.lower() in d.lower():
                    matches[term].append(os.path.relpath(os.path.join(dirpath, d), root_folder))
        # Check files
        for f in filenames:
            for term in suspicious_terms:
                if term.lower() in f.lower():
                    matches[term].append(os.path.relpath(os.path.join(dirpath, f), root_folder))

    today = datetime.now().strftime("%Y-%m-%d")
    rows = []
    for term in sorted(suspicious_terms):
        paths = matches[term]
        rows.append({
            "Suspicious Term": term,
            f"{today} (Hits)": len(paths),
            f"{today} (Paths)": "; ".join(paths) if paths else ""
        })

    df_new = pd.DataFrame(rows)

    outpath = analysis_dir / "suspicious_files.csv"
    if outpath.exists():
        df_old = pd.read_csv(outpath)
        # Merge on "Suspicious Term"
        df_all = pd.merge(df_old, df_new, on="Suspicious Term", how="outer").sort_values("Suspicious Term").reset_index(drop=True)
        df_all.fillna("", inplace=True)
    else:
        df_all = df_new
    df_all.to_csv(outpath, index=False)
    print(f"Suspicious files/folders updated: {outpath}")
    save_detailed_suspicious_paths(matches, analysis_dir, today)

def write_suspicious_summary_excel(summary_csv, detailed_csv, output_xlsx):
    """
    Creates an Excel report with:
    - 'Suspicious Summary' sheet: Each suspicious term/date with a "Click here to see the full list" hyperlink
    - 'AllPaths' sheet: All suspicious paths, filterable
    - Hyperlinks always work (to first row for the term, or to AllPaths!A1)
    """
    # Load CSVs
    summary_df = pd.read_csv(summary_csv)
    paths_df = pd.read_csv(detailed_csv)

    # Create workbook and sheets
    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "Suspicious Summary"
    ws_paths = wb.create_sheet("AllPaths")

    # Write AllPaths (detailed) first, for hyperlink targets
    ws_paths.append(list(paths_df.columns))
    for _, row in paths_df.iterrows():
        ws_paths.append(list(row.values))

    # Write summary, adding hyperlink in (Paths) column
    ws_summary.append(list(summary_df.columns))
    for i, row in summary_df.iterrows():
        row_list = list(row.values)
        # Find index of the "(Paths)" column
        paths_col_idx = None
        for idx, col in enumerate(summary_df.columns):
            if "(Paths)" in col:
                paths_col_idx = idx
                break
        if paths_col_idx is not None:
            # For the current row, set hyperlink target:
            suspicious_term = row["Suspicious Term"]
            matches = paths_df[paths_df["Suspicious Term"] == suspicious_term]
            if not matches.empty:
                row_num = matches.index[0] + 2  # +2: header + 1-based index
                cell_ref = f"AllPaths!A{row_num}"
            else:
                cell_ref = "AllPaths!A1"
            row_list[paths_col_idx] = "Click here to see the full list"
            ws_summary.append(row_list)
            cell = ws_summary.cell(row=ws_summary.max_row, column=paths_col_idx+1)
            cell.hyperlink = cell_ref
            cell.font = Font(color="0000FF", underline="single")
        else:
            ws_summary.append(row_list)

    # Auto-fit columns in both sheets
    for ws in [ws_summary, ws_paths]:
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max(15, min(max_length+2, 50))

    wb.save(output_xlsx)
    print(f"[+] Excel report with working hyperlinks saved: {output_xlsx}")

def save_detailed_suspicious_paths(matches, analysis_dir, today):
    records = []
    for term, paths in matches.items():
        for path in paths:
            records.append({
                "Suspicious Term": term,
                "Date": today,
                "File Path": path
            })
    df = pd.DataFrame(records)
    outpath = analysis_dir / "suspicious_paths.csv"
    if outpath.exists():
        df_old = pd.read_csv(outpath)
        df = pd.concat([df_old, df], ignore_index=True).drop_duplicates()
    df.to_csv(outpath, index=False)
    print(f"Detailed suspicious paths saved: {outpath}")
def keyword_content_scan(root_folder, analysis_dir, suspicious_terms):
    """
    Scan every file (regardless of extension) in the root_folder (excluding 'Analysis'),
    searching for each suspicious term (case-insensitive) in the file content.
    Save results to 'keyword_scan.csv' in the Analysis folder.
    """
    records = []
    today = datetime.now().strftime("%Y-%m-%d")

    for dirpath, _, filenames in os.walk(root_folder):
        if "Analysis" in dirpath:
            continue
        for f in filenames:
            full_path = os.path.join(dirpath, f)
            rel_path = os.path.relpath(full_path, root_folder)
            try:
                with open(full_path, "r", encoding="utf-8", errors="ignore") as file:
                    content = file.read().lower()
                    for term in suspicious_terms:
                        if term.lower() in content:
                            records.append({
                                "Suspicious Term": term,
                                "Date": today,
                                "File Path": rel_path
                            })
            except Exception:
                continue  # Ignore files that can't be read as text

    df = pd.DataFrame(records)
    outpath = analysis_dir / "keyword_scan.csv"
    if outpath.exists():
        df_old = pd.read_csv(outpath)
        df = pd.concat([df_old, df], ignore_index=True).drop_duplicates()
    df.to_csv(outpath, index=False)
    print(f"Keyword scan complete: {outpath}")

def js_analysis(root_folder, analysis_dir, suspicious_js_patterns=None):
    if suspicious_js_patterns is None:
        suspicious_js_patterns = ["eval(", "document.write", "setTimeout", "setInterval", "Function(", "window.open", "XMLHttpRequest", "fetch(", "atob(", "btoa("]

    today = datetime.now().strftime("%Y-%m-%d")
    js_files = []
    minified_count = 0
    largest_js = {"file": "", "size": 0}
    risky_js_count = 0

    for dirpath, _, filenames in os.walk(root_folder):
        if "Analysis" in dirpath:
            continue
        for f in filenames:
            if f.lower().endswith(".js"):
                full_path = os.path.join(dirpath, f)
                rel_path = os.path.relpath(full_path, root_folder)
                try:
                    size = os.path.getsize(full_path)
                    if size > largest_js["size"]:
                        largest_js = {"file": rel_path, "size": size}
                    with open(full_path, "r", encoding="utf-8", errors="ignore") as file:
                        content = file.read()
                        is_minified = ".min.js" in f.lower() or (len(content) > 0 and max(len(line) for line in content.splitlines() or [""]) > 500)
                        if is_minified:
                            minified_count += 1
                        risky = any(pattern in content for pattern in suspicious_js_patterns)
                        if risky:
                            risky_js_count += 1
                    js_files.append({
                        "JS File": rel_path,
                        "Size (KB)": f"{size/1024:.2f}",
                        "Minified": is_minified,
                        "Contains Risky Pattern": risky
                    })
                except Exception:
                    continue

    # Prepare summary row
    summary = {
        "Date": today,
        "Total JS Files": len(js_files),
        "Minified JS Files": minified_count,
        "Largest JS File": largest_js["file"],
        "Largest JS Size (KB)": f"{largest_js['size']/1024:.2f}",
        "JS Files with Risky Patterns": risky_js_count
    }

    # Save per-file JS analysis (optional, for your review)
    per_file_outpath = analysis_dir / "js_per_file.csv"
    df_js_files = pd.DataFrame(js_files)
    df_js_files.to_csv(per_file_outpath, index=False)

    # Save summary for dashboard tracking
    summary_outpath = analysis_dir / "js_analysis.csv"
    if summary_outpath.exists():
        df_old = pd.read_csv(summary_outpath)
        # If today's date already exists, replace it
        df_old = df_old[df_old["Date"] != today]
        df_all = pd.concat([df_old, pd.DataFrame([summary])], ignore_index=True)
    else:
        df_all = pd.DataFrame([summary])
    df_all.to_csv(summary_outpath, index=False)
    print(f"JS analysis complete: {summary_outpath}, per-file detail: {per_file_outpath}")

def hidden_files_folders_analysis_aggressive(root_folder, analysis_dir, suspicious_terms):
    today = datetime.now().strftime("%Y-%m-%d")

    # Well-known sensitive filenames (case-insensitive)
    sensitive_files = [
        ".git", ".env", ".htaccess", ".htpasswd", ".DS_Store", ".svn", ".bash_history", ".ssh",
        "config.php", "wp-config.php", "local.settings.json", "settings.py", "settings.json",
        "backup.sql", "db.sql", "credentials.txt", "secrets.json", "id_rsa", "id_rsa.pub",
        "docker-compose.yml", "dockerfile", ".bashrc", ".bash_profile"
    ]
    sensitive_files = [name.lower() for name in sensitive_files]
    suspicious_terms = [t.lower() for t in suspicious_terms]

    records = []

    for dirpath, dirnames, filenames in os.walk(root_folder):
        if "Analysis" in dirpath:
            continue
        # Check directories
        for d in dirnames:
            d_lc = d.lower()
            # Aggressive: flag if hidden, in sensitive list, or contains any suspicious term
            if (
                d.startswith(".")
                or d_lc in sensitive_files
                or any(term in d_lc for term in suspicious_terms)
            ):
                records.append({
                    "Date": today,
                    "Name": d,
                    "Type": "Directory",
                    "Relative Path": os.path.relpath(os.path.join(dirpath, d), root_folder)
                })
        # Check files
        for f in filenames:
            f_lc = f.lower()
            if (
                f.startswith(".")
                or f_lc in sensitive_files
                or any(term in f_lc for term in suspicious_terms)
            ):
                records.append({
                    "Date": today,
                    "Name": f,
                    "Type": "File",
                    "Relative Path": os.path.relpath(os.path.join(dirpath, f), root_folder)
                })

    df = pd.DataFrame(records)
    outpath = analysis_dir / "hidden_files.csv"
    if outpath.exists():
        df_old = pd.read_csv(outpath)
        df = pd.concat([df_old, df], ignore_index=True).drop_duplicates()
    df.to_csv(outpath, index=False)
    print(f"Aggressive hidden/sensitive files analysis complete: {outpath}")

def archive_executable_analysis(root_folder, analysis_dir):
    today = datetime.now().strftime("%Y-%m-%d")

    # Common archive and executable extensions (lowercase, without dot)
    archive_exts = ["zip", "rar", "7z", "tar", "gz", "bz2", "xz", "tgz", "tar.gz", "tar.bz2", "iso"]
    exec_exts = ["exe", "dll", "bin", "bat", "msi", "sh", "pyc", "scr", "com", "vbs", "ps1", "apk", "app", "cmd", "jar", "wsf"]

    archive_exts = set(archive_exts)
    exec_exts = set(exec_exts)

    records = []

    for dirpath, _, filenames in os.walk(root_folder):
        if "Analysis" in dirpath:
            continue
        for f in filenames:
            full_path = os.path.join(dirpath, f)
            rel_path = os.path.relpath(full_path, root_folder)
            ext = Path(f).suffix.lower().lstrip('.')
            file_type = None
            if ext in archive_exts:
                file_type = "Archive"
            elif ext in exec_exts:
                file_type = "Executable"
            if file_type:
                size = os.path.getsize(full_path)
                records.append({
                    "Date": today,
                    "File Type": file_type,
                    "Name": f,
                    "Relative Path": rel_path,
                    "Size (KB)": f"{size/1024:.2f}"
                })

    df = pd.DataFrame(records)
    outpath = analysis_dir / "archive_executables.csv"
    if outpath.exists():
        df_old = pd.read_csv(outpath)
        df = pd.concat([df_old, df], ignore_index=True).drop_duplicates()
    df.to_csv(outpath, index=False)
    print(f"Archive/Executable files analysis complete: {outpath}")

def image_file_analysis(root_folder, analysis_dir, large_size_kb=1024):
    today = datetime.now().strftime("%Y-%m-%d")
    image_exts = [
        "jpg", "jpeg", "png", "gif", "bmp", "svg", "ico", "webp", "tiff"
    ]
    image_exts = set(image_exts)

    per_file_records = []
    type_counts = {ext: 0 for ext in image_exts}
    large_images = []

    for dirpath, _, filenames in os.walk(root_folder):
        if "Analysis" in dirpath:
            continue
        for f in filenames:
            ext = Path(f).suffix.lower().lstrip('.')
            if ext in image_exts:
                full_path = os.path.join(dirpath, f)
                rel_path = os.path.relpath(full_path, root_folder)
                size_kb = os.path.getsize(full_path) / 1024
                per_file_records.append({
                    "Date": today,
                    "Image Name": f,
                    "Type": ext,
                    "Relative Path": rel_path,
                    "Size (KB)": f"{size_kb:.2f}"
                })
                type_counts[ext] += 1
                if size_kb > large_size_kb:
                    large_images.append({
                        "Date": today,
                        "Image Name": f,
                        "Relative Path": rel_path,
                        "Size (KB)": f"{size_kb:.2f}"
                    })

    # Per-file image details
    df_files = pd.DataFrame(per_file_records)
    outpath_files = analysis_dir / "image_files.csv"
    if outpath_files.exists():
        df_old = pd.read_csv(outpath_files)
        df_files = pd.concat([df_old, df_files], ignore_index=True).drop_duplicates()
    df_files.to_csv(outpath_files, index=False)

    # Summary counts
    type_counts_row = {"Date": today}
    type_counts_row.update({k.upper(): v for k, v in type_counts.items()})
    df_counts = pd.DataFrame([type_counts_row])
    outpath_counts = analysis_dir / "image_file_stats.csv"
    if outpath_counts.exists():
        df_old = pd.read_csv(outpath_counts)
        df_counts = pd.concat([df_old, df_counts], ignore_index=True).drop_duplicates()
    df_counts.to_csv(outpath_counts, index=False)

    # Large images
    df_large = pd.DataFrame(large_images)
    outpath_large = analysis_dir / "large_images.csv"
    if outpath_large.exists():
        df_old = pd.read_csv(outpath_large)
        df_large = pd.concat([df_old, df_large], ignore_index=True).drop_duplicates()
    df_large.to_csv(outpath_large, index=False)

    print(f"Image file analysis complete: {outpath_files}, {outpath_counts}, {outpath_large}")

def generate_html_dashboard(analysis_dir, output_html="analysis_dashboard.html"):
    dashboard_sections = []

    def render_table(csv_path, title):
        path = Path(analysis_dir) / csv_path
        if not path.exists():
            return f"<h3>{title}</h3><p><em>No data found.</em></p>"
        df = pd.read_csv(path)
        html = df.to_html(index=False, border=1, classes='data-table')
        return f"<h3>{title}</h3>{html}"

    # Main page header and navigation (with doubled curly braces in CSS)
    dashboard_sections.append("""
    <html>
    <head>
    <title>Automated Forensic Analysis Dashboard</title>
    <style>
        body {{ font-family: Segoe UI, Arial, sans-serif; margin: 20px; background: #f7f7fa; }}
        h2 {{ color: #38517c; }}
        h3 {{ margin-top: 36px; color: #4d4d4d; }}
        .data-table {{ border-collapse: collapse; margin-bottom: 24px; }}
        .data-table th, .data-table td {{ border: 1px solid #b8b8b8; padding: 6px 14px; }}
        .data-table th {{ background: #e0e4ed; }}
        .data-table tr:nth-child(even) {{ background: #f2f6ff; }}
        .section-link {{ margin-right: 18px; }}
        a {{ color: #2066a2; text-decoration: none; }}
        a:hover {{ text-decoration: underline; }}
    </style>
    </head>
    <body>
    <h2>📊 Automated Forensic Analysis Dashboard</h2>
    <p>Last updated: <b>{}</b></p>
    <p>
      <a class="section-link" href="#filetypes">File Types</a>
      <a class="section-link" href="#top10">Top 10 Largest Files</a>
      <a class="section-link" href="#js">JavaScript Analysis</a>
      <a class="section-link" href="#archives">Archives & Executables</a>
      <a class="section-link" href="#images">Image Stats</a>
      <a class="section-link" href="#suspicious">Suspicious Files/Names</a>
      <a class="section-link" href="#hidden">Hidden Files/Folders</a>
    </p>
    """.format(pd.Timestamp.now().strftime("%d/%m/%Y %H:%M")))

    # Add tables for each module
    dashboard_sections.append('<a name="filetypes"></a>' + render_table("file_type_stats.csv", "File Type Stats"))
    dashboard_sections.append('<a name="top10"></a>' + render_table("top10_largest_files.csv", "Top 10 Largest Files"))
    dashboard_sections.append('<a name="js"></a>' + render_table("js_analysis.csv", "JavaScript Analysis Summary"))
    dashboard_sections.append('<a name="archives"></a>' + render_table("archive_executables.csv", "Archives & Executables"))
    dashboard_sections.append('<a name="images"></a>' + render_table("image_file_stats.csv", "Image File Stats"))
    dashboard_sections.append('<a name="suspicious"></a>' + render_table("suspicious_files.csv", "Suspicious Files/Names"))
    dashboard_sections.append('<a name="hidden"></a>' + render_table("hidden_files.csv", "Hidden/Sensitive Files & Folders"))

    dashboard_sections.append("""
    <br>
    <p style="font-size:14px;color:#888;">Download full datasets:
        <a href="file_type_stats.csv" download>file_type_stats.csv</a> |
        <a href="top10_largest_files.csv" download>top10_largest_files.csv</a> |
        <a href="js_analysis.csv" download>js_analysis.csv</a> |
        <a href="archive_executables.csv" download>archive_executables.csv</a> |
        <a href="image_file_stats.csv" download>image_file_stats.csv</a> |
        <a href="suspicious_files.csv" download>suspicious_files.csv</a> |
        <a href="hidden_files.csv" download>hidden_files.csv</a>
    </p>
    </body></html>
    """)

    dashboard = "\n".join(dashboard_sections)
    outpath = Path(analysis_dir) / output_html
    with open(outpath, "w", encoding="utf-8") as f:
        f.write(dashboard)
    print(f"HTML dashboard generated: {outpath}")

if __name__ == "__main__":
    base_folder = Path('.').resolve()
    analysis_dir = ensure_analysis_folder(base_folder)
    update_file_type_stats(base_folder, analysis_dir)
    update_top10_largest_files(base_folder, analysis_dir)
    update_suspicious_files_folders(base_folder, analysis_dir)
    keyword_content_scan(base_folder, analysis_dir, suspicious_terms)
    js_analysis(base_folder, analysis_dir)
    hidden_files_folders_analysis_aggressive(base_folder, analysis_dir, suspicious_terms)
    archive_executable_analysis(base_folder, analysis_dir)
    image_file_analysis(base_folder, analysis_dir, large_size_kb=1024)  # 1024KB = 1MB, adjust as needed
    generate_html_dashboard(analysis_dir)
    # write_suspicious_summary_excel(
    # str(analysis_dir / "suspicious_files.csv"),
    # str(analysis_dir / "suspicious_paths.csv"),
    # str(analysis_dir / "suspicious_report.xlsx")
#)